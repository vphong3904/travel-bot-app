# app/api/routes/admin.py
"""
Admin routes — /api/admin/*

Tất cả endpoints yêu cầu role admin. Ma trận quyền theo nghiệp vụ:
- moderator       : chỉ giám sát (đọc) Users/Chat/Feedback — MONITOR_ROLES.
- content_manager : chỉ nhập liệu Content/KB/Media/City Mapping/Intent Patterns
                     /Locations/Tours/Itineraries — CONTENT_ROLES.
- admin           : kế thừa moderator + content_manager, thêm quyền quản lý user
                     (khoá tài khoản, tạo/xoá) + vận hành nội bộ (RAG monitor,
                     session) — STAFF_ROLES. KHÔNG backup, KHÔNG system-config,
                     KHÔNG đổi role người khác.
- super_admin     : toàn quyền, gồm backup, system-config, đổi role.
Xem app/api/deps.py (MONITOR_ROLES / CONTENT_ROLES / STAFF_ROLES).
"""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File, status
from sqlalchemy import cast, func, or_, select, update, delete, Date, case
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import (
    DB, CurrentUser, ADMIN_ROLES, MONITOR_ROLES, CONTENT_ROLES, STAFF_ROLES,
    require_admin, require_role,
)
from app.db.database import AsyncSessionLocal
from app.db.models.admin import EmbeddingJob, KnowledgeEntry
from app.db.models.media import ContentItem, ContentOption
from app.db.models.chat import ChatMessage, ChatSession
from app.db.models.travel import Destination, Ticket, Review, City
from app.db.models.user import User
from app.services.knowledge import KnowledgeService
from app.core.security import hash_password
from app.utils import get_logger
from app.utils.image_processing import process_image

logger = get_logger("admin")
router = APIRouter(tags=["admin"])

# ── Guard: mọi endpoint đều cần admin role ────────────────────────────────────

def _admin(user: CurrentUser) -> User:
    if user.role not in ADMIN_ROLES:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Admin access required")
    return user

AdminUser = CurrentUser  # re-use; guard applied per endpoint with Depends


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD STATS
# ══════════════════════════════════════════════════════════════════════════════

async def _overview_data(db: AsyncSession, period: str) -> dict:
    """Dashboard tổng quan — trả về đúng cấu trúc DashboardOverview cho Flutter."""
    now = datetime.now(timezone.utc)
    delta = {"day": 1, "week": 7, "month": 30, "year": 365}[period]
    since = now - timedelta(days=delta)

    # ── KPI counts ────────────────────────────────────────────────────────────
    total_users = await db.scalar(
        select(func.count(User.id)).where(User.is_deleted == False)
    )
    new_users = await db.scalar(
        select(func.count(User.id)).where(
            User.created_at >= since, User.is_deleted == False
        )
    )
    total_sessions = await db.scalar(select(func.count(ChatSession.id)))

    # Chỉ đếm assistant messages để tính answered_rate
    total_messages = await db.scalar(select(func.count(ChatMessage.id)))
    assistant_messages = await db.scalar(
        select(func.count(ChatMessage.id)).where(ChatMessage.role == "assistant")
    )
    user_messages = await db.scalar(
        select(func.count(ChatMessage.id)).where(ChatMessage.role == "user")
    )

    # answered_rate = số assistant msg / số user msg (tối đa 1.0)
    answered_rate: float = 0.0
    if user_messages and user_messages > 0:
        answered_rate = min(1.0, (assistant_messages or 0) / user_messages)

    # Pending: session bị flag / messages không có câu trả lời (heuristic)
    pending_flagged = await db.scalar(
        select(func.count(ChatSession.id)).where(ChatSession.is_flagged == True)
    )
    # Unanswered = user messages có feedback âm hoặc chưa có feedback
    pending_unanswered = await db.scalar(
        select(func.count(ChatMessage.id)).where(
            ChatMessage.role == "user",
            ChatMessage.feedback == None,  # noqa: E711
            ChatMessage.created_at >= since,
        )
    )

    # ── Time-series: users_over_time ─────────────────────────────────────────
    users_ts_rows = await db.execute(
        select(
            cast(func.date_trunc("day", User.created_at), Date).label("date"),
            func.count(User.id).label("count"),
        )
        .where(User.created_at >= since, User.is_deleted == False)
        .group_by("date")
        .order_by("date")
    )
    users_over_time = [
        {"date": str(r.date), "count": int(r.count)}
        for r in users_ts_rows
    ]

    # ── Time-series: messages_over_time ──────────────────────────────────────
    messages_ts_rows = await db.execute(
        select(
            cast(func.date_trunc("day", ChatMessage.created_at), Date).label("date"),
            func.count(ChatMessage.id).label("count"),
        )
        .where(ChatMessage.created_at >= since)
        .group_by("date")
        .order_by("date")
    )
    messages_over_time = [
        {"date": str(r.date), "count": int(r.count)}
        for r in messages_ts_rows
    ]

    # ── Top destinations (by favorite_count + view_count) ────────────────────
    top_dest_rows = await db.execute(
        select(Destination.name, Destination.favorite_count, Destination.view_count)
        .where(Destination.is_active == True)
        .order_by((Destination.favorite_count + Destination.view_count).desc())
        .limit(10)
    )
    top_destinations = [
        {"destination": r.name, "count": int((r.favorite_count or 0) + (r.view_count or 0))}
        for r in top_dest_rows
    ]

    # ── Intent breakdown (từ ChatMessage.intent trong period) ─────────────────
    intent_rows = await db.execute(
        select(
            ChatMessage.intent.label("intent"),
            func.count(ChatMessage.id).label("count"),
        )
        .where(
            ChatMessage.intent != None,  # noqa: E711
            ChatMessage.created_at >= since,
        )
        .group_by(ChatMessage.intent)
        .order_by(func.count(ChatMessage.id).desc())
        .limit(10)
    )
    intent_breakdown = [
        {"intent": r.intent, "count": int(r.count)}
        for r in intent_rows
    ]

    return {
        "period": period,
        "kpi": {
            "total_users": total_users or 0,
            "new_users_this_period": new_users or 0,
            "total_chat_sessions": total_sessions or 0,
            "total_messages": total_messages or 0,
            "answered_rate": round(answered_rate, 4),
            "pending_unanswered": pending_unanswered or 0,
            "pending_flagged": pending_flagged or 0,
        },
        "users_over_time": users_over_time,
        "messages_over_time": messages_over_time,
        "top_destinations": top_destinations,
        "intent_breakdown": intent_breakdown,
    }


@router.get("/stats/overview")
async def stats_overview(
    period: str = Query("month", pattern="^(day|week|month|year)$"),
    db: DB = None,
    _: User = Depends(require_admin),
):
    return await _overview_data(db, period)


@router.get("/stats/feedback")
async def stats_feedback(
    period: str = Query("month", pattern="^(day|week|month|year)$"),
    db: DB = None,
    _: User = Depends(require_role(MONITOR_ROLES)),
):
    """Thống kê feedback theo ngày trong period."""
    now = datetime.now(timezone.utc)
    delta = {"day": 1, "week": 7, "month": 30, "year": 365}[period]
    since = now - timedelta(days=delta)

    rows = await db.execute(
        select(
            func.date_trunc("day", ChatMessage.created_at).label("date"),
            func.sum(
                case((ChatMessage.feedback == 1, 1), else_=0)
            ).label("positive"),
            func.sum(
                case((ChatMessage.feedback == -1, 1), else_=0)
            ).label("negative"),
        )
        .where(ChatMessage.created_at >= since)
        .group_by("date")
        .order_by("date")
    )
    daily = [
        {
            "date": str(r.date)[:10],
            "positive": int(r.positive or 0),
            "negative": int(r.negative or 0),
        }
        for r in rows
    ]
    return {"daily": daily, "period": period}


@router.get("/stats/export")
async def stats_export(
    format: str = Query("excel"),
    report: str = Query("overview"),
    period: str = Query("month", pattern="^(day|week|month|year)$"),
    db: DB = None,
    _: User = Depends(require_admin),
):
    from io import BytesIO
    from fastapi.responses import Response
    from openpyxl import Workbook

    data = await _overview_data(db, period)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tổng quan"
    ws.append(["Chỉ số", "Giá trị"])
    kpi_labels = {
        "total_users": "Tổng người dùng",
        "new_users_this_period": "Người dùng mới",
        "total_chat_sessions": "Tổng phiên chat",
        "total_messages": "Tổng tin nhắn",
        "answered_rate": "Tỉ lệ trả lời",
        "pending_unanswered": "Chưa trả lời",
        "pending_flagged": "Bị gắn cờ",
    }
    for key, label in kpi_labels.items():
        ws.append([label, data["kpi"].get(key)])

    ws2 = wb.create_sheet("Người dùng theo ngày")
    ws2.append(["Ngày", "Số lượng"])
    for row in data["users_over_time"]:
        ws2.append([row["date"], row["count"]])

    ws3 = wb.create_sheet("Tin nhắn theo ngày")
    ws3.append(["Ngày", "Số lượng"])
    for row in data["messages_over_time"]:
        ws3.append([row["date"], row["count"]])

    ws4 = wb.create_sheet("Điểm đến nổi bật")
    ws4.append(["Điểm đến", "Lượt quan tâm"])
    for row in data["top_destinations"]:
        ws4.append([row["destination"], row["count"]])

    ws5 = wb.create_sheet("Intent")
    ws5.append(["Intent", "Số lượng"])
    for row in data["intent_breakdown"]:
        ws5.append([row["intent"], row["count"]])

    buf = BytesIO()
    wb.save(buf)
    filename = f"pdtrip_{report}_{period}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ══════════════════════════════════════════════════════════════════════════════
# USER MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/users")
async def list_users(
    q: Optional[str] = None,
    role: Optional[str] = None,
    is_active: Optional[bool] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, le=100),
    db: DB = None,
    _: User = Depends(require_role(STAFF_ROLES)),
):
    stmt = select(User).where(User.is_deleted == False)
    if q:
        stmt = stmt.where(
            User.email.ilike(f"%{q}%")
            | User.username.ilike(f"%{q}%")
            | User.full_name.ilike(f"%{q}%")
        )
    if role:
        stmt = stmt.where(User.role == role)
    if is_active is not None:
        stmt = stmt.where(User.is_active == is_active)

    total = await db.scalar(select(func.count()).select_from(stmt.subquery()))
    rows = await db.execute(stmt.offset((page - 1) * limit).limit(limit))
    users = rows.scalars().all()

    return {
        "total": total or 0,
        "page": page,
        "items": [_user_dict(u) for u in users],
    }


@router.post("/users", status_code=201)
async def create_user(
    body: dict,
    db: DB = None,
    actor: User = Depends(require_role(["admin", "super_admin"])),
):
    username = (body.get("username") or "").strip()
    email = (body.get("email") or "").strip().lower()
    password = body.get("password") or ""
    role = body.get("role", "user")
    if not username or not email or not password:
        raise HTTPException(400, "Thiếu username/email/password")
    if len(password) < 8:
        raise HTTPException(400, "Mật khẩu tối thiểu 8 ký tự")
    if role == "super_admin" and actor.role != "super_admin":
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Chỉ Super Admin mới được tạo tài khoản Super Admin",
        )

    existing = await db.execute(
        select(User).where(
            (User.email == email) | (User.username == username)
        ).limit(1)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(409, "Email hoặc username đã tồn tại")

    u = User(
        username=username,
        email=email,
        password_hash=hash_password(password),
        full_name=body.get("full_name"),
        role=role,
        auth_provider="email",
    )
    db.add(u)
    await db.commit()
    await db.refresh(u)
    return _user_dict(u)


@router.get("/users/{user_id}")
async def get_user(
    user_id: str,
    db: DB = None,
    _: User = Depends(require_role(STAFF_ROLES)),
):
    u = await _get_or_404(db, User, user_id)

    sessions_count = await db.scalar(
        select(func.count(ChatSession.id)).where(
            ChatSession.user_id == user_id, ChatSession.is_deleted == False
        )
    )
    messages_count = await db.scalar(
        select(func.count(ChatMessage.id)).join(
            ChatSession, ChatMessage.session_id == ChatSession.id
        ).where(ChatSession.user_id == user_id)
    )

    recent_sessions_rows = await db.execute(
        select(ChatSession)
        .where(ChatSession.user_id == user_id, ChatSession.is_deleted == False)
        .order_by(ChatSession.updated_at.desc())
        .limit(5)
    )
    recent_sessions = [
        {
            "id": s.id,
            "title": s.title,
            "total_messages": s.total_messages,
            "updated_at": str(s.updated_at),
        }
        for s in recent_sessions_rows.scalars()
    ]

    return {
        **_user_dict(u),
        "total_chat_sessions": sessions_count or 0,
        "total_messages": messages_count or 0,
        "recent_sessions": recent_sessions,
    }


@router.patch("/users/{user_id}")
async def update_user(
    user_id: str,
    body: dict,
    db: DB = None,
    actor: User = Depends(require_role(STAFF_ROLES)),
):
    u = await _get_or_404(db, User, user_id)
    if u.role == "super_admin" and actor.role != "super_admin":
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Chỉ Super Admin mới được chỉnh sửa tài khoản Super Admin",
        )
    if "is_active" in body:
        u.is_active = bool(body["is_active"])
    if "role" in body:
        if actor.role not in ("admin", "super_admin"):
            raise HTTPException(status.HTTP_403_FORBIDDEN, "Chỉ admin mới đổi được role")
        if body["role"] == "super_admin" and actor.role != "super_admin":
            raise HTTPException(
                status.HTTP_403_FORBIDDEN,
                "Chỉ Super Admin mới được thay đổi vai trò liên quan đến Super Admin",
            )
        u.role = body["role"]
    if "full_name" in body:
        u.full_name = body["full_name"]
    u.updated_at = datetime.now(timezone.utc)
    await db.commit()
    return _user_dict(u)


@router.patch("/users/{user_id}/role")
async def change_role(
    user_id: str,
    body: dict,
    db: DB = None,
    actor: User = Depends(require_role(["super_admin"])),
):
    """Chỉ Super Admin được cấp/đổi vai trò — Admin không được quyền này."""
    u = await _get_or_404(db, User, user_id)
    new_role = body.get("role", u.role)
    u.role = new_role
    u.updated_at = datetime.now(timezone.utc)
    await db.commit()
    return _user_dict(u)


@router.delete("/users/{user_id}")
async def delete_user(
    user_id: str,
    db: DB = None,
    actor: User = Depends(require_role(["admin", "super_admin"])),
):
    if user_id == actor.id:
        raise HTTPException(400, "Không thể tự xoá chính mình")
    u = await _get_or_404(db, User, user_id)
    if u.role == "super_admin" and actor.role != "super_admin":
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Chỉ Super Admin mới được xoá tài khoản Super Admin",
        )
    u.is_deleted = True
    u.is_active = False
    u.updated_at = datetime.now(timezone.utc)
    await db.commit()
    return {"ok": True, "id": user_id}


# ══════════════════════════════════════════════════════════════════════════════
# REVIEWS (moderation) — admin thấy tất cả review, chỉ admin/super_admin được
# xoá. Xem MONITOR_ROLES (moderator giám sát), xoá STAFF_ROLES (admin trở lên) —
# khớp quy ước RBAC ở đầu file: moderator chỉ đọc, không sửa.
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/reviews")
async def list_reviews_admin(
    search: Optional[str] = None,
    destination_id: Optional[UUID] = None,
    rating: Optional[int] = Query(None, ge=1, le=5),
    skip: int = 0,
    limit: int = Query(20, le=100),
    db: DB = None,
    _: User = Depends(require_role(MONITOR_ROLES)),
):
    """Danh sách review toàn hệ thống để admin duyệt/kiểm soát nội dung phá hoại."""
    stmt = (
        select(Review, User.username, User.email, User.full_name,
               Destination.name.label("destination_name"))
        .join(User, Review.user_id == User.id)
        .join(Destination, Review.destination_id == Destination.id)
    )
    if search:
        like = f"%{search}%"
        stmt = stmt.where(or_(
            Review.content.ilike(like),
            User.username.ilike(like),
            User.email.ilike(like),
        ))
    if destination_id:
        stmt = stmt.where(Review.destination_id == destination_id)
    if rating:
        stmt = stmt.where(Review.rating == rating)

    total = await db.scalar(select(func.count()).select_from(stmt.subquery()))
    rows = (await db.execute(
        stmt.order_by(Review.created_at.desc()).offset(skip).limit(limit)
    )).all()

    items = [
        {
            "id": str(review.id),
            "user_id": str(review.user_id),
            "username": username,
            "user_email": email,
            "user_full_name": full_name,
            "destination_id": str(review.destination_id),
            "destination_name": destination_name,
            "rating": review.rating,
            "content": review.content,
            "created_at": review.created_at.isoformat() if review.created_at else None,
        }
        for review, username, email, full_name, destination_name in rows
    ]
    return {"items": items, "total": total or 0}


@router.delete("/reviews/{review_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_review_admin(
    review_id: UUID,
    db: DB = None,
    _: User = Depends(require_role(STAFF_ROLES)),
):
    """Xoá review bất kỳ (moderation) — chỉ admin/super_admin, tái tính rating
    ngay sau khi xoá (trigger DB chỉ xử lý INSERT)."""
    review = (await db.execute(
        select(Review).where(Review.id == review_id)
    )).scalar_one_or_none()
    if not review:
        raise HTTPException(status_code=404, detail="Review không tồn tại")

    destination_id = review.destination_id
    await db.delete(review)
    await db.commit()

    from app.api.routes.reviews import _recalculate_review_stats
    await _recalculate_review_stats(db, destination_id)


# ══════════════════════════════════════════════════════════════════════════════
# CHAT SESSIONS (admin view)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/chat-sessions")
async def list_chat_sessions(
    user_id: Optional[str] = None,
    is_flagged: Optional[bool] = None,
    q: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, le=100),
    db: DB = None,
    _: User = Depends(require_role(MONITOR_ROLES)),
):
    stmt = select(ChatSession).where(ChatSession.is_deleted == False)
    if user_id:
        stmt = stmt.where(ChatSession.user_id == user_id)
    if is_flagged is not None:
        stmt = stmt.where(ChatSession.is_flagged == is_flagged)
    if q:
        stmt = stmt.where(ChatSession.title.ilike(f"%{q}%"))

    total = await db.scalar(select(func.count()).select_from(stmt.subquery()))
    rows = await db.execute(
        stmt.order_by(ChatSession.updated_at.desc()).offset((page - 1) * limit).limit(limit)
    )
    sessions = rows.scalars().all()

    return {
        "total": total or 0,
        "page": page,
        "items": [
            {
                "id": s.id,
                "user_id": s.user_id,
                "title": s.title,
                "total_messages": s.total_messages,
                "is_flagged": s.is_flagged,
                "created_at": str(s.created_at),
                "updated_at": str(s.updated_at),
            }
            for s in sessions
        ],
    }


@router.get("/chat-sessions/{session_id}/messages")
async def get_session_messages(
    session_id: str,
    db: DB = None,
    _: User = Depends(require_role(MONITOR_ROLES)),
):
    # FE (chat_view) cần {session:{...}, messages:[...]} — không phải list thô.
    session = await db.get(ChatSession, session_id)
    if not session:
        raise HTTPException(404, "Session không tồn tại")
    rows = await db.execute(
        select(ChatMessage)
        .where(ChatMessage.session_id == session_id)
        .order_by(ChatMessage.created_at)
    )
    messages = rows.scalars().all()
    return {
        "session": {
            "id": session.id,
            "user_id": session.user_id,
            "title": session.title,
            "total_messages": session.total_messages,
            "is_flagged": session.is_flagged,
            "tags": getattr(session, "tags", None) or [],
            "created_at": str(session.created_at),
            "updated_at": str(session.updated_at),
        },
        "messages": [
            {
                "id": m.id,
                "role": m.role,
                "content": m.content,
                "intent": m.intent,
                "feedback": m.feedback,
                "latency_ms": m.latency_ms,
                "created_at": str(m.created_at),
            }
            for m in messages
        ],
    }


@router.patch("/chat-sessions/{session_id}")
async def update_chat_session(
    session_id: str,
    body: dict,
    db: DB = None,
    _: User = Depends(require_role(STAFF_ROLES)),
):
    """Admin cập nhật cờ đánh dấu (is_flagged) và/hoặc tags của 1 hội thoại."""
    session = await db.get(ChatSession, session_id)
    if not session:
        raise HTTPException(404, "Session không tồn tại")

    if body.get("is_flagged") is not None:
        session.is_flagged = bool(body["is_flagged"])
    if body.get("tags") is not None:
        session.tags = [str(t) for t in body["tags"]]
    session.updated_at = datetime.now(timezone.utc)

    await db.commit()
    return {
        "ok": True,
        "id": session.id,
        "is_flagged": session.is_flagged,
        "tags": session.tags or [],
    }


# ══════════════════════════════════════════════════════════════════════════════
# KNOWLEDGE BASE
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/knowledge")
async def list_knowledge(
    q: Optional[str] = None,
    category: Optional[str] = None,
    is_active: Optional[bool] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, le=100),
    db: DB = None,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    stmt = select(KnowledgeEntry)
    if q:
        stmt = stmt.where(
            KnowledgeEntry.title.ilike(f"%{q}%") | KnowledgeEntry.content.ilike(f"%{q}%")
        )
    if category:
        stmt = stmt.where(KnowledgeEntry.category == category)
    if is_active is not None:
        stmt = stmt.where(KnowledgeEntry.is_active == is_active)

    total = await db.scalar(select(func.count()).select_from(stmt.subquery()))
    rows = await db.execute(
        stmt.order_by(KnowledgeEntry.created_at.desc()).offset((page - 1) * limit).limit(limit)
    )
    entries = rows.scalars().all()

    return {
        "total": total or 0,
        "page": page,
        "items": [
            {
                "id": e.id,
                "title": e.title,
                "content": e.content or "",
                "category": e.category,
                "tags": e.tags or [],
                "is_active": e.is_active,
                "qdrant_id": e.qdrant_id,
                "embedding_status": "done" if e.qdrant_id else "not_embedded",
                "source": e.source,
                "created_at": str(e.created_at),
                "updated_at": str(e.updated_at),
            }
            for e in entries
        ],
    }


@router.post("/knowledge", status_code=201)
async def create_knowledge(
    body: dict,
    db: DB = None,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    svc = KnowledgeService(db)
    entry = await svc.create(
        title=body["title"],
        category=body["category"],
        content=body["content"],
        tags=body.get("tags", []),
        source=body.get("source"),
    )
    return {"id": entry.id, "title": entry.title, "status": "pending"}


@router.patch("/knowledge/{entry_id}")
async def update_knowledge(
    entry_id: str,
    body: dict,
    db: DB = None,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    svc = KnowledgeService(db)
    entry = await svc.update(
        entry_id=entry_id,
        title=body.get("title"),
        content=body.get("content"),
        tags=body.get("tags"),
        is_active=body.get("is_active"),
    )
    return {"id": entry.id, "title": entry.title, "status": "pending"}


@router.delete("/knowledge/{entry_id}")
async def delete_knowledge(
    entry_id: str,
    db: DB = None,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    svc = KnowledgeService(db)
    await svc.delete(entry_id)
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# EMBEDDING JOBS
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/embedding-jobs/{job_id}")
async def get_embedding_job(
    job_id: str,
    db: DB = None,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    result = await db.execute(
        select(EmbeddingJob).where(EmbeddingJob.id == job_id)
    )
    job = result.scalar_one_or_none()
    if not job:
        raise HTTPException(404, "Job not found")
    return {
        "id": job.id,
        "entity_type": job.entity_type,
        "entity_id": job.entity_id,
        "status": job.status,
        "error": job.error,
        "created_at": str(job.created_at),
        "updated_at": str(job.updated_at),
    }


# ══════════════════════════════════════════════════════════════════════════════
# KB HEALTH CHECK
# ══════════════════════════════════════════════════════════════════════════════

# knowledge-base/{city_slug}/{file} — loại nội dung -> tên file trên đĩa.
# Khớp đúng _ctLabels ở frontend (kb_health_screen.dart).
_KB_CONTENT_TYPES: list[tuple[str, str]] = [
    ("destinations", "destinations.json"),
    ("hotels", "hotels.json"),
    ("restaurants", "restaurants.json"),
    ("foods", "foods.json"),
    ("transport", "transport.json"),
    ("tours", "tours.json"),
    ("events", "events.json"),
    ("shopping", "shopping.json"),
    ("itineraries", "itineraries.json"),
    ("experiences", "experiences.md"),
    ("faq", "faq.md"),
]

# backend/app/api/routes/admin.py -> backend/knowledge-base
_KB_ROOT = Path(__file__).resolve().parent.parent.parent.parent / "knowledge-base"

# File nhỏ hơn ngưỡng này coi như "rỗng" (chỉ có khung sẵn, chưa nhập liệu thật).
_KB_EMPTY_SIZE_THRESHOLD = 200


def _kb_file_has_data(path: Path, size: int) -> bool:
    if size < _KB_EMPTY_SIZE_THRESHOLD:
        return False
    if path.suffix == ".json":
        try:
            content = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return False
        if isinstance(content, (list, dict)):
            return len(content) > 0
        return bool(content)
    return True  # .md — đã qua ngưỡng kích thước, coi là có nội dung


@router.get("/kb-health")
async def kb_health(
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    """
    Quét knowledge-base/ trên đĩa: mỗi thành phố (folder) x mỗi loại nội dung
    (destinations/hotels/.../faq) → có file chưa, file có dữ liệu thật không.
    Dùng để admin biết nhanh thành phố nào còn thiếu nội dung cần nhập liệu.

    [Fix] Bản cũ trả về thống kê Postgres (KnowledgeEntry/EmbeddingJob) + debug
    Qdrant — hoàn toàn KHÔNG khớp với contract mà frontend
    (kb_health_repository.dart -> KbHealthResponse.fromJson) đã thiết kế sẵn
    (summary/content_types/cities theo shape quét file này), nên FE luôn crash
    lúc parse response ("Lỗi: type 'Null' is not a subtype of type
    'Map<String, dynamic>'"). Không có consumer nào khác dùng shape cũ.
    """
    cities: list[dict] = []
    if _KB_ROOT.is_dir():
        for city_dir in sorted(_KB_ROOT.iterdir()):
            if not city_dir.is_dir():
                continue
            files: dict[str, dict] = {}
            filled = 0
            for ct, filename in _KB_CONTENT_TYPES:
                fpath = city_dir / filename
                exists = fpath.is_file()
                size = fpath.stat().st_size if exists else 0
                has_data = exists and _kb_file_has_data(fpath, size)
                if has_data:
                    filled += 1
                files[ct] = {
                    "exists": exists,
                    "has_data": has_data,
                    "size_bytes": size if exists else None,
                    "last_modified": (
                        datetime.fromtimestamp(fpath.stat().st_mtime, tz=timezone.utc).isoformat()
                        if exists else None
                    ),
                }
            total = len(_KB_CONTENT_TYPES)
            cities.append({
                "city_slug": city_dir.name,
                "filled_count": filled,
                "total_count": total,
                "completeness_pct": round(filled / total * 100) if total else 0,
                "has_any_data": filled > 0,
                "files": files,
            })

    total_cities = len(cities)
    complete_cities = sum(1 for c in cities if c["completeness_pct"] == 100)
    empty_cities = sum(1 for c in cities if not c["has_any_data"])
    avg_completeness = (
        round(sum(c["completeness_pct"] for c in cities) / total_cities)
        if total_cities else 0
    )

    return {
        "summary": {
            "total_cities": total_cities,
            "complete_cities": complete_cities,
            "empty_cities": empty_cities,
            "avg_completeness_pct": avg_completeness,
        },
        "content_types": [ct for ct, _ in _KB_CONTENT_TYPES],
        "cities": cities,
    }


# ══════════════════════════════════════════════════════════════════════════════
# RAG MONITORING
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/rag-monitoring/overview")
async def rag_overview(
    period: str = Query("week"),
    date_from: Optional[str] = Query(None, alias="from"),
    date_to: Optional[str] = Query(None, alias="to"),
    db: DB = None,
    _: User = Depends(require_role(STAFF_ROLES)),
):
    def _parse_utc(s: str) -> datetime:
        dt = datetime.fromisoformat(s)
        return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)

    # Ưu tiên khoảng thời gian tuỳ chọn (from/to) nếu FE gửi lên; ngược lại
    # dùng bucket cố định theo `period` (day/week/month).
    until = datetime.now(timezone.utc)
    if date_from:
        since = _parse_utc(date_from)
        if date_to:
            until = _parse_utc(date_to)
    else:
        delta = {"day": 1, "week": 7, "month": 30}.get(period, 7)
        since = until - timedelta(days=delta)

    avg_confidence = await db.scalar(
        select(func.avg(ChatMessage.confidence_score)).where(
            ChatMessage.created_at >= since, ChatMessage.created_at <= until,
            ChatMessage.confidence_score.isnot(None)
        )
    )
    avg_search_ms = await db.scalar(
        select(func.avg(ChatMessage.search_ms)).where(
            ChatMessage.created_at >= since, ChatMessage.created_at <= until,
            ChatMessage.search_ms.isnot(None)
        )
    )
    avg_llm_ms = await db.scalar(
        select(func.avg(ChatMessage.llm_ms)).where(
            ChatMessage.created_at >= since, ChatMessage.created_at <= until,
            ChatMessage.llm_ms.isnot(None)
        )
    )
    avg_chunk_count = await db.scalar(
        select(func.avg(ChatMessage.chunk_count)).where(
            ChatMessage.created_at >= since, ChatMessage.created_at <= until,
            ChatMessage.chunk_count.isnot(None)
        )
    )

    # hallucination_rate = tỉ lệ messages có confidence < 0.3
    total_assistant = await db.scalar(
        select(func.count(ChatMessage.id)).where(
            ChatMessage.role == "assistant",
            ChatMessage.created_at >= since,
            ChatMessage.created_at <= until,
            ChatMessage.confidence_score.isnot(None),
        )
    ) or 0
    low_confidence = await db.scalar(
        select(func.count(ChatMessage.id)).where(
            ChatMessage.role == "assistant",
            ChatMessage.created_at >= since,
            ChatMessage.created_at <= until,
            ChatMessage.confidence_score < 0.3,
        )
    ) or 0
    hallucination_rate = (low_confidence / total_assistant) if total_assistant > 0 else 0.0

    # cache_hit_rate breakdown
    cache_rows = await db.execute(
        select(ChatMessage.cache_hit, func.count(ChatMessage.id).label("cnt"))
        .where(
            ChatMessage.created_at >= since, ChatMessage.created_at <= until,
            ChatMessage.cache_hit.isnot(None),
        )
        .group_by(ChatMessage.cache_hit)
    )
    total_cache = 0
    cache_counts: dict = {}
    for r in cache_rows:
        cache_counts[r.cache_hit] = r.cnt
        total_cache += r.cnt
    cache_hit_rate = {
        k: round(v / total_cache, 3) if total_cache > 0 else 0.0
        for k, v in cache_counts.items()
    }

    # search_method breakdown
    method_rows = await db.execute(
        select(ChatMessage.search_method, func.count(ChatMessage.id).label("cnt"))
        .where(
            ChatMessage.created_at >= since, ChatMessage.created_at <= until,
            ChatMessage.search_method.isnot(None),
        )
        .group_by(ChatMessage.search_method)
    )
    total_methods = 0
    method_counts: dict = {}
    for r in method_rows:
        method_counts[r.search_method] = r.cnt
        total_methods += r.cnt
    search_method_breakdown = {
        k: round(v / total_methods, 3) if total_methods > 0 else 0.0
        for k, v in method_counts.items()
    }

    # confidence_over_time
    conf_rows = await db.execute(
        select(
            cast(func.date_trunc("day", ChatMessage.created_at), Date).label("date"),
            func.avg(ChatMessage.confidence_score).label("avg_score"),
        )
        .where(
            ChatMessage.created_at >= since, ChatMessage.created_at <= until,
            ChatMessage.confidence_score.isnot(None),
        )
        .group_by("date")
        .order_by("date")
    )
    confidence_over_time = [
        {"date": str(r.date), "avg_score": round(float(r.avg_score or 0), 3)}
        for r in conf_rows
    ]

    return {
        "period": period,
        "avg_confidence_score": round(float(avg_confidence or 0), 3),
        "avg_search_ms": round(float(avg_search_ms or 0), 1),
        "avg_llm_ms": round(float(avg_llm_ms or 0), 1),
        "avg_chunk_count": round(float(avg_chunk_count or 0), 1),
        "hallucination_rate": round(hallucination_rate, 3),
        "cache_hit_rate": cache_hit_rate,
        "search_method_breakdown": search_method_breakdown,
        "confidence_over_time": confidence_over_time,
    }


@router.get("/rag-monitoring/latency")
async def rag_latency(
    period: str = Query("week"),
    db: DB = None,
    _: User = Depends(require_role(STAFF_ROLES)),
):
    delta = {"day": 1, "week": 7, "month": 30}.get(period, 7)
    since = datetime.now(timezone.utc) - timedelta(days=delta)

    rows = await db.execute(
        select(
            func.date_trunc("day", ChatMessage.created_at).label("date"),
            func.avg(ChatMessage.latency_ms).label("avg_ms"),
            func.percentile_cont(0.95).within_group(ChatMessage.latency_ms).label("p95_ms"),
        )
        .where(ChatMessage.created_at >= since, ChatMessage.latency_ms.isnot(None))
        .group_by("date")
        .order_by("date")
    )
    return {
        "items": [
            {
                "date": str(r.date)[:10],
                "avg_ms": round(float(r.avg_ms or 0), 1),
                "p95_ms": round(float(r.p95_ms or 0), 1),
            }
            for r in rows
        ]
    }


@router.get("/rag-monitoring/errors")
async def rag_errors(
    period: str = Query("week"),
    db: DB = None,
    _: User = Depends(require_role(STAFF_ROLES)),
):
    delta = {"day": 1, "week": 7, "month": 30}.get(period, 7)
    since = datetime.now(timezone.utc) - timedelta(days=delta)

    rows = await db.execute(
        select(ChatMessage)
        .where(
            ChatMessage.created_at >= since,
            ChatMessage.role == "assistant",
            ChatMessage.confidence_score < 0.3,
        )
        .order_by(ChatMessage.created_at.desc())
        .limit(50)
    )
    messages = rows.scalars().all()
    return {
        "items": [
            {
                "id": m.id,
                "content": m.content[:200],
                "confidence_score": m.confidence_score,
                "intent": m.intent,
                "created_at": str(m.created_at),
            }
            for m in messages
        ]
    }


# ══════════════════════════════════════════════════════════════════════════════
# CITY MAPPINGS
# ══════════════════════════════════════════════════════════════════════════════

# Simple in-memory city mappings (can be moved to DB later)
_city_overrides: dict[str, str] = {}


@router.get("/city-mappings/validate")
async def validate_city_mappings(
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    """Trả danh sách province names và mapping hiện tại."""
    from app.services.nlp_preprocessor import INTENT_PATTERNS
    import os

    kb_path = "knowledge-base"
    existing_folders = set()
    if os.path.isdir(kb_path):
        existing_folders = {
            d for d in os.listdir(kb_path)
            if os.path.isdir(os.path.join(kb_path, d))
        }

    # Lấy danh sách provinces từ DB
    return [
        {
            "old_province": slug,
            "mapped_slug": _city_overrides.get(slug, slug),
            "folder_exists": _city_overrides.get(slug, slug) in existing_folders,
            "suggestion": None,
        }
        for slug in list(existing_folders)[:50]
    ]


@router.get("/city-mappings/valid-slugs")
async def valid_slugs(
    db: DB = None,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    """Trả danh sách city slug từ bảng destinations."""
    from app.db.models.travel import Destination
    rows = await db.execute(
        select(Destination.slug)
        .where(Destination.slug.isnot(None), Destination.is_active == True)
        .order_by(Destination.slug)
    )
    slugs = [r[0] for r in rows if r[0]]
    return slugs or []


@router.patch("/city-mappings/{old_province}")
async def update_city_mapping(
    old_province: str,
    body: dict,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    _city_overrides[old_province] = body["new_slug"]
    return {"ok": True, "old_province": old_province, "new_slug": body["new_slug"]}


# ══════════════════════════════════════════════════════════════════════════════
# INTENT PATTERNS
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/intent-patterns")
async def list_intent_patterns(
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    from app.services.nlp_preprocessor import INTENT_PATTERNS
    result: dict = {}
    for intent, kws in INTENT_PATTERNS.items():
        keyword_list = [
            {
                "keyword": kw,
                "is_collision": sum(1 for _, v in INTENT_PATTERNS.items() if kw in v) > 1,
            }
            for kw in kws
        ]
        result[intent] = {"keywords": keyword_list, "collision_warnings": []}
    return result


@router.post("/intent-patterns/{intent}/keywords", status_code=201)
async def add_keyword(
    intent: str,
    body: dict,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    from app.services.nlp_preprocessor import INTENT_PATTERNS, reload_intent_patterns
    kw = body.get("keyword", "").strip().lower()
    if not kw:
        raise HTTPException(400, "keyword không được trống")
    if intent not in INTENT_PATTERNS:
        raise HTTPException(404, f"Intent '{intent}' không tồn tại")
    if kw not in INTENT_PATTERNS[intent]:
        INTENT_PATTERNS[intent].append(kw)
        _save_intent_patterns(INTENT_PATTERNS)
        reload_intent_patterns()
    return {"ok": True}


@router.delete("/intent-patterns/{intent}/keywords/{keyword}")
async def delete_keyword(
    intent: str,
    keyword: str,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    from app.services.nlp_preprocessor import INTENT_PATTERNS, reload_intent_patterns
    if intent not in INTENT_PATTERNS:
        raise HTTPException(404, f"Intent '{intent}' không tồn tại")
    if keyword in INTENT_PATTERNS[intent]:
        INTENT_PATTERNS[intent].remove(keyword)
        _save_intent_patterns(INTENT_PATTERNS)
        reload_intent_patterns()
    return {"ok": True}


@router.post("/intent-patterns/test")
async def test_intent(body: dict, _: User = Depends(require_role(CONTENT_ROLES))):
    from app.services.nlp_preprocessor import detect_intent, extract_entities
    text = body.get("text", "")
    entities = extract_entities(text)
    intent, confidence = detect_intent(text, entities)
    return {"intent": intent, "confidence": confidence, "matched_keywords": []}


def _save_intent_patterns(patterns: dict) -> None:
    """Lưu patterns về file JSON."""
    import os
    path = os.path.join(os.path.dirname(__file__), "..", "..", "..", "app", "data", "intent_patterns.json")
    path = os.path.normpath(path)
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(patterns, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"Không thể lưu intent_patterns.json: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# CONTENT MANAGEMENT (generic CRUD)
# ══════════════════════════════════════════════════════════════════════════════

# content_type hợp lệ. faq & experiences KHÔNG ở đây — chúng là knowledge entries
# (quản lý qua /admin/knowledge). Content lưu ở bảng content_items (JSONB).
CONTENT_TYPES: set[str] = {
    "destinations", "hotels", "tours", "foods", "restaurants",
    "shopping", "itineraries", "events", "transport",
}


def _serialize_content(c: ContentItem) -> dict:
    """Chuẩn hoá 1 ContentItem cho FE (ContentItem.fromJson)."""
    data = dict(c.data or {})
    data.setdefault("name", c.name)
    if c.image_url:
        data["image_url"] = c.image_url
    return {
        "id": str(c.id),
        "status": c.status,
        "is_deleted": c.is_deleted,
        "image_url": c.image_url,
        "created_at": str(c.created_at),
        "updated_at": str(c.updated_at),
        "data": data,
    }


def _extract_content_fields(body: dict) -> tuple[str, str | None, dict]:
    """Tách name / image_url ra khỏi phần data động."""
    data = {k: v for k, v in body.items() if k not in ("city_slug",)}
    name = (data.get("name") or "Untitled").strip() or "Untitled"
    image_url = data.get("image_url") or None
    return name, image_url, data


@router.get("/content/{content_type}")
async def list_content(
    content_type: str,
    city_slug: Optional[str] = None,       # optional — không còn "cổng bắt buộc chọn city"
    status: Optional[str] = None,
    search: Optional[str] = None,
    sort: str = Query("newest"),           # newest | oldest | name
    date_from: Optional[str] = None,       # YYYY-MM-DD
    date_to: Optional[str] = None,         # YYYY-MM-DD
    field: Optional[str] = None,           # key trong data để lọc
    value: Optional[str] = None,           # giá trị khớp (ilike)
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: DB = None,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    if content_type not in CONTENT_TYPES:
        raise HTTPException(400, f"Content type '{content_type}' không hợp lệ")

    stmt = select(ContentItem).where(
        ContentItem.content_type == content_type,
        ContentItem.is_deleted == False,  # noqa: E712
    )
    if city_slug:
        stmt = stmt.where(ContentItem.city_slug == city_slug)
    if status:
        stmt = stmt.where(ContentItem.status == status)
    if search:
        stmt = stmt.where(ContentItem.name.ilike(f"%{search}%"))
    if date_from:
        stmt = stmt.where(cast(ContentItem.created_at, Date) >= date_from)
    if date_to:
        stmt = stmt.where(cast(ContentItem.created_at, Date) <= date_to)
    if field and value:
        stmt = stmt.where(ContentItem.data[field].astext.ilike(f"%{value}%"))

    total = await db.scalar(select(func.count()).select_from(stmt.subquery()))
    order = {
        "oldest": ContentItem.created_at.asc(),
        "name": ContentItem.name.asc(),
    }.get(sort, ContentItem.created_at.desc())
    rows = await db.execute(
        stmt.order_by(order).offset((page - 1) * page_size).limit(page_size)
    )
    items = rows.scalars().all()
    return {
        "total": total or 0,
        "page": page,
        "items": [_serialize_content(c) for c in items],
    }


@router.post("/content/{content_type}", status_code=201)
async def create_content(
    content_type: str,
    body: dict,
    city_slug: Optional[str] = None,
    db: DB = None,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    if content_type not in CONTENT_TYPES:
        raise HTTPException(400, "Content type không hợp lệ")
    name, image_url, data = _extract_content_fields(body)
    item = ContentItem(
        content_type=content_type,
        city_slug=city_slug or body.get("city_slug"),
        name=name,
        data=data,
        image_url=image_url,
        status="draft",
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return _serialize_content(item)


@router.patch("/content/{content_type}/{item_id}")
async def update_content(
    content_type: str,
    item_id: str,
    body: dict,
    db: DB = None,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    item = await db.get(ContentItem, item_id)
    if not item or item.is_deleted:
        raise HTTPException(404, "Không tìm thấy mục")
    name, image_url, data = _extract_content_fields(body)
    item.name = name
    item.data = data
    item.image_url = image_url
    item.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(item)
    return _serialize_content(item)


@router.delete("/content/{content_type}/{item_id}")
async def delete_content(
    content_type: str,
    item_id: str,
    db: DB = None,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    item = await db.get(ContentItem, item_id)
    if item:
        item.is_deleted = True
        item.updated_at = datetime.now(timezone.utc)
        await db.commit()
    return {"ok": True}


@router.patch("/content/{content_type}/{item_id}/publish")
async def publish_content(
    content_type: str,
    item_id: str,
    db: DB = None,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    await db.execute(
        update(ContentItem)
        .where(ContentItem.id == item_id)
        .values(status="published", updated_at=datetime.now(timezone.utc))
    )
    await db.commit()
    return {"ok": True}


# ── Cities (master list cho dropdown/filter content) ──────────────────────────

@router.get("/cities")
async def list_cities(
    q: Optional[str] = None,
    db: DB = None,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    """
    Danh sách city (mức điểm đến) cho dropdown filter. Search KHÔNG phân biệt dấu
    (unaccent) theo tên điểm đến, tên tỉnh mới (34), hoặc alias tỉnh cũ (63) —
    gõ "da lat" ra Đà Lạt, "kien giang" ra Phú Quốc.
    """
    stmt = select(City).where(City.is_active == True)  # noqa: E712
    if q:
        like = f"%{q}%"
        stmt = stmt.where(
            func.unaccent(City.name).ilike(func.unaccent(like))
            | func.unaccent(func.coalesce(City.province, "")).ilike(func.unaccent(like))
            | func.unaccent(func.array_to_string(City.old_aliases, " ")).ilike(
                func.unaccent(like))
        )
    rows = await db.execute(stmt.order_by(City.province, City.name))
    return [
        {"id": str(c.id), "slug": c.slug, "name": c.name,
         "province": c.province, "region": c.region}
        for c in rows.scalars().all()
    ]


# ── Content options (taxonomy: "loại" theo content_type + field) ──────────────

def _serialize_option(o: ContentOption) -> dict:
    return {
        "id": str(o.id), "content_type": o.content_type, "field": o.field,
        "code": o.code, "label": o.label, "sort_order": o.sort_order,
        "is_active": o.is_active,
    }


@router.get("/content-options")
async def list_content_options(
    content_type: Optional[str] = None,
    field: Optional[str] = None,
    db: DB = None,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    """Danh sách options. Form lọc theo content_type+field (và is_active phía FE);
    màn quản lý lấy tất cả."""
    stmt = select(ContentOption)
    if content_type:
        stmt = stmt.where(ContentOption.content_type == content_type)
    if field:
        stmt = stmt.where(ContentOption.field == field)
    rows = await db.execute(
        stmt.order_by(ContentOption.content_type, ContentOption.field,
                      ContentOption.sort_order, ContentOption.label)
    )
    return [_serialize_option(o) for o in rows.scalars().all()]


@router.post("/content-options", status_code=201)
async def create_content_option(
    body: dict, db: DB = None, _: User = Depends(require_role(CONTENT_ROLES)),
):
    for f in ("content_type", "field", "code", "label"):
        if not body.get(f):
            raise HTTPException(422, f"Thiếu '{f}'")
    obj = ContentOption(
        content_type=body["content_type"], field=body["field"],
        code=body["code"].strip(), label=body["label"].strip(),
        sort_order=body.get("sort_order", 0),
    )
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return _serialize_option(obj)


@router.patch("/content-options/{opt_id}")
async def update_content_option(
    opt_id: str, body: dict, db: DB = None, _: User = Depends(require_role(CONTENT_ROLES)),
):
    obj = await _get_or_404(db, ContentOption, opt_id)
    for f in ("content_type", "field", "code", "label", "sort_order", "is_active"):
        if f in body:
            setattr(obj, f, body[f])
    await db.commit()
    return _serialize_option(obj)


@router.delete("/content-options/{opt_id}", status_code=204)
async def delete_content_option(
    opt_id: str, db: DB = None, _: User = Depends(require_role(CONTENT_ROLES)),
):
    obj = await _get_or_404(db, ContentOption, opt_id)
    await db.delete(obj)
    await db.commit()


# ══════════════════════════════════════════════════════════════════════════════
# CHATBOT TEST (admin thử chatbot ngay trong panel — không lưu DB, không giới hạn)
# ══════════════════════════════════════════════════════════════════════════════

_admin_rag = None


def _clean_bot_answer(text: str) -> str:
    """Bỏ markdown ** và trích dẫn [1] khỏi câu trả lời chatbot cho gọn."""
    text = text or ""
    text = re.sub(r"\s*\[\d+\]", "", text)      # bỏ [1], [2]... (kèm space trước)
    text = text.replace("**", "")               # bỏ in đậm markdown
    text = re.sub(r"(?m)^\s*\*\s+", "• ", text)  # bullet '*   ' → '• '
    return text.strip()


def _get_admin_rag():
    global _admin_rag
    if _admin_rag is None:
        from app.services.rag_pipeline import RAGPipeline
        _admin_rag = RAGPipeline()
    return _admin_rag


@router.post("/chatbot/test")
async def chatbot_test(
    body: dict,
    _: User = Depends(require_admin),
):
    """
    Chạy RAG cho 1 câu hỏi để admin test chatbot. KHÔNG lưu session/history vào DB,
    không giới hạn. `history` (tùy chọn) = [{"role","content"}, ...] cho ngữ cảnh.
    """
    question = (body.get("content") or "").strip()
    if not question:
        raise HTTPException(422, "content không được trống")
    history = body.get("history") or []
    rag = _get_admin_rag()
    result = await rag.query(
        question=question,
        history=history,
        session_id="admin-test",
    )
    return {
        "answer": _clean_bot_answer(result.get("answer", "")),
        "intent": result.get("intent"),
        "confidence_score": result.get("confidence_score"),
        "sources": result.get("sources", []),
        "latency_ms": result.get("latency_ms"),
        "search_method": result.get("search_method"),
    }


# ══════════════════════════════════════════════════════════════════════════════
# FEEDBACK MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/feedback")
async def list_feedback(
    type: Optional[str] = None,
    category: Optional[str] = None,
    intent: Optional[str] = None,
    page: int = Query(1, ge=1),
    db: DB = None,
    _: User = Depends(require_role(MONITOR_ROLES)),
):
    stmt = select(ChatMessage).where(ChatMessage.feedback.isnot(None))
    if type == "positive":
        stmt = stmt.where(ChatMessage.feedback == 1)
    elif type == "negative":
        stmt = stmt.where(ChatMessage.feedback == -1)
    if intent:
        stmt = stmt.where(ChatMessage.intent == intent)
    if category:
        stmt = stmt.where(ChatMessage.feedback_category == category)

    total = await db.scalar(select(func.count()).select_from(stmt.subquery()))
    rows = await db.execute(
        stmt.order_by(ChatMessage.created_at.desc()).offset((page - 1) * 20).limit(20)
    )
    messages = rows.scalars().all()

    return {
        "total": total or 0,
        "page": page,
        "items": [
            {
                "message_id": m.id,
                "session_id": m.session_id,
                "content_preview": (m.content or "")[:150],
                "feedback_type": "positive" if m.feedback == 1 else "negative",
                # đúng key mà FeedbackItem.fromJson đọc + giá trị THẬT từ DB
                "feedback_category": m.feedback_category,
                "feedback_reason": m.feedback_reason,
                "intent": m.intent,
                "feedback_resolved": bool(m.feedback_resolved),
                "created_at": str(m.created_at),
            }
            for m in messages
        ],
    }


@router.patch("/feedback/{message_id}/resolve")
async def resolve_feedback(
    message_id: str,
    db: DB = None,
    _: User = Depends(require_role(STAFF_ROLES)),
):
    # feedback_resolved column được thêm ở migration 28
    try:
        await db.execute(
            update(ChatMessage)
            .where(ChatMessage.id == message_id)
            .values(feedback_resolved=True)
        )
        await db.commit()
    except Exception:
        # Column chưa tồn tại nếu migration chưa chạy
        pass
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# MEDIA — trình quản lý ảnh dạng CMS (thư mục + ảnh, lưu DB + local disk)
# ══════════════════════════════════════════════════════════════════════════════
#
# - media_folders: cây thư mục (root / folder con) — TA-018 + yêu cầu CMS.
# - media_files:  ảnh, gắn folder_id, soft delete. File lưu static/uploads,
#                 mount /uploads ở main.py. URL = /uploads/<filename>.
# - Upload: multi-select nhiều ảnh 1 lần, Pillow resize ≤1920 → WebP (best-effort).
# Role: đọc + tạo/sửa/xoá = admin/super_admin/content_manager (CONTENT_ROLES).
# Moderator không thuộc nhóm nội dung nên không xem được media.

import os as _os

_UPLOAD_DIR = _os.path.join("static", "uploads")
_ALLOWED_EXT = {"jpg", "jpeg", "png", "webp"}
_MAX_UPLOAD_BYTES = 8 * 1024 * 1024  # giới hạn input (sau resize WebP còn nhỏ hơn)

_media_writer = require_role(CONTENT_ROLES)


def _process_image(content: bytes, ext: str) -> tuple[bytes, str, Optional[str], Optional[int], Optional[int]]:
    return process_image(content, ext, max_side=1920)


def _folder_uuid(value: Optional[str]) -> Optional[UUID]:
    if not value or str(value).lower() in ("", "null", "none", "root"):
        return None
    try:
        return UUID(str(value))
    except (ValueError, TypeError):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "folder_id không hợp lệ")


# ── Folders ───────────────────────────────────────────────────────────────────

@router.get("/media/folders")
async def list_media_folders(db: DB = None, _: User = Depends(require_role(CONTENT_ROLES))):
    """Trả về toàn bộ thư mục (phẳng) kèm số ảnh + thời điểm thêm ảnh gần nhất.

    FE tự dựng cây từ parent_id và sắp xếp thư mục theo last_added để hiển thị
    thư mục vừa thêm ảnh gần nhất lên trước.
    """
    from app.db.models.media import MediaFolder, MediaFile

    folders = (await db.execute(
        select(MediaFolder).order_by(MediaFolder.created_at.asc())
    )).scalars().all()

    # Đếm ảnh + last_added theo folder (1 query gom nhóm)
    stats = (await db.execute(
        select(
            MediaFile.folder_id,
            func.count(MediaFile.id),
            func.max(MediaFile.created_at),
        )
        .where(MediaFile.is_deleted == False)
        .group_by(MediaFile.folder_id)
    )).all()
    by_folder = {str(fid): (cnt, last) for fid, cnt, last in stats if fid is not None}

    return [
        {
            "id": str(f.id),
            "name": f.name,
            "parent_id": str(f.parent_id) if f.parent_id else None,
            "image_count": by_folder.get(str(f.id), (0, None))[0],
            "last_added": (by_folder.get(str(f.id), (0, None))[1].isoformat()
                           if by_folder.get(str(f.id), (0, None))[1] else None),
            "created_at": f.created_at.isoformat() if f.created_at else None,
        }
        for f in folders
    ]


@router.post("/media/folders", status_code=201)
async def create_media_folder(
    body: dict, db: DB = None, actor: User = Depends(_media_writer),
):
    from app.db.models.media import MediaFolder

    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Tên thư mục không được trống")
    parent_id = _folder_uuid(body.get("parent_id"))

    if parent_id is not None:
        parent = (await db.execute(
            select(MediaFolder).where(MediaFolder.id == parent_id)
        )).scalar_one_or_none()
        if not parent:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Thư mục cha không tồn tại")

    folder = MediaFolder(name=name, parent_id=parent_id, created_by=actor.id)
    db.add(folder)
    try:
        await db.commit()
    except Exception:
        await db.rollback()
        raise HTTPException(status.HTTP_409_CONFLICT, "Đã có thư mục trùng tên trong thư mục này")
    await db.refresh(folder)
    return {
        "id": str(folder.id),
        "name": folder.name,
        "parent_id": str(folder.parent_id) if folder.parent_id else None,
        "image_count": 0,
        "last_added": None,
        "created_at": folder.created_at.isoformat() if folder.created_at else None,
    }


@router.patch("/media/folders/{folder_id}")
async def rename_media_folder(
    folder_id: str, body: dict, db: DB = None, _: User = Depends(_media_writer),
):
    from app.db.models.media import MediaFolder

    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Tên thư mục không được trống")
    folder = (await db.execute(
        select(MediaFolder).where(MediaFolder.id == _folder_uuid(folder_id))
    )).scalar_one_or_none()
    if not folder:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Thư mục không tồn tại")
    folder.name = name
    try:
        await db.commit()
    except Exception:
        await db.rollback()
        raise HTTPException(status.HTTP_409_CONFLICT, "Đã có thư mục trùng tên trong thư mục này")
    return {"id": folder_id, "name": name}


@router.delete("/media/folders/{folder_id}")
async def delete_media_folder(
    folder_id: str, db: DB = None, _: User = Depends(_media_writer),
):
    """Xoá thư mục (cascade thư mục con); ảnh bên trong gỡ folder_id (giữ ảnh)."""
    from app.db.models.media import MediaFolder

    fid = _folder_uuid(folder_id)
    folder = (await db.execute(
        select(MediaFolder).where(MediaFolder.id == fid)
    )).scalar_one_or_none()
    if not folder:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Thư mục không tồn tại")
    await db.delete(folder)
    await db.commit()
    return {"ok": True}


# ── Files ─────────────────────────────────────────────────────────────────────

@router.get("/media")
async def list_media(
    folder_id: Optional[str] = None,
    tag: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(24, ge=1, le=100),
    db: DB = None,
    _: User = Depends(require_role(CONTENT_ROLES)),
):
    from app.db.models.media import MediaFile

    stmt = select(MediaFile).where(MediaFile.is_deleted == False)
    if folder_id is not None:
        stmt = stmt.where(MediaFile.folder_id == _folder_uuid(folder_id))
    if tag:
        stmt = stmt.where(MediaFile.tags.any(tag))

    total = (await db.execute(
        select(func.count()).select_from(stmt.subquery())
    )).scalar() or 0

    rows = (await db.execute(
        stmt.order_by(MediaFile.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )).scalars().all()

    return {
        "total": total,
        "page": page,
        "items": [
            {
                "id": str(m.id),
                "filename": m.filename,
                "original_name": m.original_name,
                "file_path": m.file_path,
                "file_size": m.file_size,
                "mime_type": m.mime_type,
                "width": m.width,
                "height": m.height,
                "tags": list(m.tags or []),
                "folder_id": str(m.folder_id) if m.folder_id else None,
                "created_at": m.created_at.isoformat() if m.created_at else None,
            }
            for m in rows
        ],
    }


@router.post("/media/upload", status_code=201)
async def upload_media(
    files: list[UploadFile] = File(...),
    folder_id: Optional[str] = Query(None),
    db: DB = None,
    actor: User = Depends(_media_writer),
):
    """Upload nhiều ảnh 1 lần vào thư mục đang mở (folder_id)."""
    from app.db.models.media import MediaFile, MediaFolder
    import uuid as _uuid

    fid = _folder_uuid(folder_id)
    if fid is not None:
        exists = (await db.execute(
            select(MediaFolder.id).where(MediaFolder.id == fid)
        )).scalar_one_or_none()
        if not exists:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Thư mục không tồn tại")

    _os.makedirs(_UPLOAD_DIR, exist_ok=True)
    saved = []
    for file in files:
        ext = (_os.path.splitext(file.filename or "")[1].lstrip(".") or "jpg").lower()
        if ext not in _ALLOWED_EXT:
            raise HTTPException(status.HTTP_400_BAD_REQUEST,
                                f"Định dạng không hỗ trợ: {file.filename} (chỉ jpg/png/webp)")
        content = await file.read()
        if len(content) > _MAX_UPLOAD_BYTES:
            raise HTTPException(status.HTTP_400_BAD_REQUEST,
                                f"Ảnh quá lớn (>8MB): {file.filename}")

        out, out_ext, mime, width, height = _process_image(content, ext)
        filename = f"{_uuid.uuid4()}.{out_ext}"
        with open(_os.path.join(_UPLOAD_DIR, filename), "wb") as f:
            f.write(out)

        record = MediaFile(
            filename=filename,
            original_name=file.filename,
            file_path=f"/uploads/{filename}",
            file_size=len(out),
            mime_type=mime or file.content_type,
            width=width,
            height=height,
            tags=[],
            folder_id=fid,
            uploaded_by=actor.id,
        )
        db.add(record)
        saved.append(record)

    await db.commit()
    for r in saved:
        await db.refresh(r)
    return [
        {
            "id": str(r.id),
            "filename": r.filename,
            "original_name": r.original_name,
            "file_path": r.file_path,
            "file_size": r.file_size,
            "mime_type": r.mime_type,
            "width": r.width,
            "height": r.height,
            "tags": [],
            "folder_id": str(r.folder_id) if r.folder_id else None,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in saved
    ]


@router.delete("/media/{media_id}")
async def delete_media(
    media_id: str, db: DB = None, _: User = Depends(_media_writer),
):
    from app.db.models.media import MediaFile

    try:
        mid = UUID(media_id)
    except (ValueError, TypeError):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "media_id không hợp lệ")
    await db.execute(
        update(MediaFile).where(MediaFile.id == mid).values(is_deleted=True)
    )
    await db.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# UNANSWERED QUESTIONS
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/unanswered-questions")
async def list_unanswered(
    db: DB = None,
    _: User = Depends(require_role(MONITOR_ROLES)),
):
    rows = await db.execute(
        select(ChatMessage)
        .where(
            ChatMessage.role == "assistant",
            ChatMessage.confidence_score < 0.3,
        )
        .order_by(ChatMessage.created_at.desc())
        .limit(50)
    )
    messages = rows.scalars().all()
    # FE (unanswered_list) mong {items:[...]} với field `question` + `is_promoted`.
    return {
        "total": len(messages),
        "items": [
            {
                "id": m.id,
                "question": (m.content or "")[:300],
                "confidence_score": m.confidence_score,
                "intent": m.intent,
                "is_promoted": False,
                "created_at": str(m.created_at),
            }
            for m in messages
        ],
    }


@router.post("/unanswered-questions/{question_id}/promote-to-kb")
async def promote_to_kb(
    question_id: str,
    body: dict | None = None,
    db: DB = None,
    _: User = Depends(require_role(STAFF_ROLES)),
):
    # FE có thể gọi không kèm body → tự lấy title/content từ chính message.
    body = body or {}
    title = body.get("title")
    content = body.get("content")
    if not title or not content:
        msg = await db.get(ChatMessage, question_id)
        text = (msg.content if msg else "") or ""
        title = title or (text[:80] or "Câu hỏi chưa trả lời")
        content = content or text
    svc = KnowledgeService(db)
    entry = await svc.create(
        title=title,
        category=body.get("category", "faq"),
        content=content or title,
    )
    return {"id": entry.id, "status": "pending"}


@router.post("/unanswered-questions/{question_id}/ai-suggest")
async def ai_suggest_kb(
    question_id: str,
    db: DB = None,
    _: User = Depends(require_role(STAFF_ROLES)),
):
    """
    TP-005: AI (Gemini) soạn draft KB từ câu hỏi chưa trả lời được.
    Chỉ trả draft — admin sửa rồi bấm promote-to-kb (endpoint cũ) để lưu.
    """
    msg = await db.get(ChatMessage, question_id)
    if not msg:
        raise HTTPException(404, "Không tìm thấy câu hỏi")
    question = (msg.content or "").strip()
    if not question:
        raise HTTPException(400, "Câu hỏi rỗng")

    from app.services.kb_suggestion_service import suggest_kb_draft
    try:
        draft = await suggest_kb_draft(question)
    except RuntimeError as e:
        raise HTTPException(502, str(e))
    return {"question": question[:300], "draft": draft}


# ══════════════════════════════════════════════════════════════════════════════
# ANALYTICS — TOP QUESTIONS (TP-004)
# ══════════════════════════════════════════════════════════════════════════════

# Câu smalltalk phổ biến — loại khỏi thống kê "câu hỏi hay hỏi"
_SMALLTALK_QUESTIONS = (
    "xin chào", "chào", "chào bạn", "hello", "hi", "alo", "test", "ok", "oke",
    "cảm ơn", "cám ơn", "cảm ơn bạn", "thanks", "thank you", "tạm biệt", "bye",
)


@router.get("/analytics/top-questions")
async def top_questions(
    period: str = Query("week"),
    limit: int = Query(20, ge=1, le=100),
    db: DB = None,
    _: User = Depends(require_role(MONITOR_ROLES)),
):
    """Câu hỏi user hay hỏi nhất — normalize lower + trim + cắt 120 ký tự."""
    delta = {"day": 1, "week": 7, "month": 30}.get(period, 7)
    since = datetime.now(timezone.utc) - timedelta(days=delta)

    normalized = func.lower(func.left(func.trim(ChatMessage.content), 120))
    rows = await db.execute(
        select(
            normalized.label("question"),
            func.count(ChatMessage.id).label("count"),
            func.max(ChatMessage.created_at).label("last_asked"),
        )
        .where(
            ChatMessage.role == "user",
            ChatMessage.created_at >= since,
            func.length(func.trim(ChatMessage.content)) >= 8,
            normalized.notin_(_SMALLTALK_QUESTIONS),
        )
        .group_by(normalized)
        .order_by(func.count(ChatMessage.id).desc(), func.max(ChatMessage.created_at).desc())
        .limit(limit)
    )
    return {
        "period": period,
        "items": [
            {
                "question": r.question,
                "count": int(r.count),
                "last_asked": str(r.last_asked),
            }
            for r in rows
        ],
    }


# ══════════════════════════════════════════════════════════════════════════════
# SYSTEM CONFIG
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/system-config")
async def get_system_config(
    db: DB = None,
    _: User = Depends(require_role(["super_admin"])),
):
    try:
        from sqlalchemy import text
        rows = await db.execute(text("SELECT key, value, description, updated_at FROM system_configs"))
        return [
            {
                "key": r.key,
                "value": r.value,
                "description": r.description,
                "updated_at": str(r.updated_at) if r.updated_at else None,
            }
            for r in rows
        ]
    except Exception:
        return []


@router.patch("/system-config/{key}")
async def update_system_config(
    key: str,
    body: dict,
    db: DB = None,
    actor: User = Depends(require_role(["super_admin"])),
):
    from sqlalchemy import text
    value = body.get("value")
    await db.execute(
        text(
            "UPDATE system_configs SET value = :value, updated_by = :uid, updated_at = NOW() "
            "WHERE key = :key"
        ),
        {"value": json.dumps(value), "uid": actor.id, "key": key},
    )
    await db.commit()
    return {"ok": True, "key": key, "value": value}


# ══════════════════════════════════════════════════════════════════════════════
# DATABASE BACKUP
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/system/backup")
async def trigger_backup(
    _: User = Depends(require_role(["super_admin"])),
):
    """Backup thủ công — pg_dump toàn bộ DB ra backups/. Dump chứa password_hash
    nên chỉ Super Admin được phép chạy/tải."""
    from app.services import backup_service
    try:
        filepath = await backup_service.create_backup()
    except Exception as e:
        raise HTTPException(500, f"Backup thất bại: {e}")
    stat = filepath.stat()
    return {
        "filename": filepath.name,
        "size_bytes": stat.st_size,
        "created_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
    }


@router.get("/system/backups")
async def list_backups(
    _: User = Depends(require_role(["super_admin"])),
):
    from app.services import backup_service
    return {"items": backup_service.list_backups()}


@router.get("/system/backups/{filename}/download")
async def download_backup(
    filename: str,
    _: User = Depends(require_role(["super_admin"])),
):
    from fastapi.responses import FileResponse
    from app.services import backup_service

    if "/" in filename or "\\" in filename or not filename.endswith(".sql"):
        raise HTTPException(400, "Tên file không hợp lệ")
    filepath = backup_service.BACKUP_DIR / filename
    if not filepath.is_file():
        raise HTTPException(404, "Không tìm thấy file backup")
    return FileResponse(filepath, media_type="application/sql", filename=filename)


# ══════════════════════════════════════════════════════════════════════════════
# SESSION MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/sessions")
async def list_sessions(
    user_id: Optional[str] = None,
    db: DB = None,
    _: User = Depends(require_role(STAFF_ROLES)),
):
    from app.db.models.user import RefreshToken
    stmt = select(RefreshToken).where(
        RefreshToken.revoked == False,
        RefreshToken.expires_at > datetime.now(timezone.utc),
    )
    if user_id:
        stmt = stmt.where(RefreshToken.user_id == user_id)
    rows = await db.execute(stmt.order_by(RefreshToken.created_at.desc()).limit(50))
    tokens = rows.scalars().all()
    return [
        {
            "id": t.id,
            "user_id": t.user_id,
            "ip_address": None,
            "user_agent": None,
            "created_at": str(t.created_at),
            "expires_at": str(t.expires_at),
        }
        for t in tokens
    ]


@router.delete("/sessions/{session_id}")
async def revoke_session(
    session_id: str,
    db: DB = None,
    _: User = Depends(require_role(STAFF_ROLES)),
):
    from app.db.models.user import RefreshToken
    await db.execute(
        update(RefreshToken).where(RefreshToken.id == session_id).values(revoked=True)
    )
    await db.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

async def _get_or_404(db: AsyncSession, model, id: str):
    result = await db.execute(select(model).where(model.id == id))
    obj = result.scalar_one_or_none()
    if not obj:
        raise HTTPException(404, f"{model.__name__} không tồn tại")
    return obj


def _user_dict(u: User) -> dict:
    return {
        "id": u.id,
        "username": u.username,
        "email": u.email,
        "full_name": u.full_name,
        "avatar_url": u.avatar_url,
        "role": u.role,
        "is_active": u.is_active,
        "auth_provider": u.auth_provider,
        "created_at": str(u.created_at),
        "updated_at": str(u.updated_at),
    }

